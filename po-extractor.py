import os
import re
import glob
import datetime
import shutil
from tkinter import Tk, Button, Text, font, END, filedialog
from threading import Thread
from openpyxl import load_workbook
from po_formats import text_extract, po_base, po_type_1, po_type_2, po_type_3, po_type_4, po_type_5, po_type_6, po_type_7, po_type_8

class PO_Extractor:
    """
        PO Extractor GUI
    """
    def __init__(self) -> None:

        self.__root = Tk()

        # Application properties
        self.__aspectRatio = (16,9)
        self.__sizeFactor = 70
        self.__windowSize = (
            self.__aspectRatio[0]*self.__sizeFactor,
            self.__aspectRatio[1]*self.__sizeFactor
            )

        # Application Theme
        self.__fonts = (
            font.Font(family="Courier New",size=10, weight='normal'),
            font.Font(family="Courier New",size=12, weight='normal')
            )
        self.__colors = (
            '#DDE6ED',  #Light
            '#9DB2BF',
            '#526D82',
            '#27374D'   #Dark
            )

        self.__root.title("PO Extractor")
        self.__root.geometry(f"{self.__windowSize[0]}x{self.__windowSize[1]}")
        self.__root.resizable(False, False)
        self.__root.configure(background=self.__colors[1])
        self.__srcDir = "C:/PDFextract"
        self.__poFilesList = []
        self.__buildLayout()

    def __buildLayout(self)->None:
        """
            Create the tkinter GUI layout of the application
        """
        self.__createButton_selectDirectory()
        self.__createButton_extractData()
        self.__createButton_saveLog()
        self.__createTextField_log()

    def __createButton_selectDirectory(self)->None:
        """
            Create the tkinter button for 'Select Directory'
        """
        self.__buttonSelect = Button(
            self.__root, 
            text="Select Source Directory",
            font=self.__fonts[1],
            bg = self.__colors[0],
            fg = self.__colors[3],
            activebackground=self.__colors[1],
            width=44,
            relief='groove',
            command=self.__selectDirectory
            )
        self.__buttonSelect.grid(
            row=0,column=0,
            ipadx=10,ipady=10,
            padx=10,pady=10)

    def __createButton_extractData(self)->None:
        """
            Create the tkinter button for 'Extract PO Data'
        """
        def startThread():
            _t = Thread(target=self.__startExtraction)
            _t.start()
        self.__buttonExtract = Button(
            self.__root, 
            text="Extract PO Data",
            font=self.__fonts[1],
            bg = self.__colors[0],
            fg = self.__colors[3],
            activebackground=self.__colors[1],
            width=44,
            relief='groove',
            command=startThread
            )
        self.__buttonExtract.grid(
            row=0,column=1,
            ipadx=10,ipady=10,
            padx=10,pady=10)

    def __createButton_saveLog(self)->None:
        self.__buttonSaveLog = Button(
            self.__root, 
            text="Save Log",
            font=self.__fonts[1],
            bg = self.__colors[0],
            fg = self.__colors[3],
            activebackground=self.__colors[1],
            width=8,
            relief='groove',
            command=self.__saveLog
            )
        self.__buttonSaveLog.grid(
            row=0,column=2,
            ipadx=10,ipady=10,
            padx=10,pady=10)

    def __createTextField_log(self)->None:
        """
            Create the tkinter text field for log printing
        """
        self.__textFieldLog = Text(
            self.__root, 
            font=self.__fonts[0],
            bg = self.__colors[0],
            fg = self.__colors[3],
            width=134,
            height=32,
            relief='flat',
            state='disabled'
            )
        self.__textFieldLog.grid(
            row=1,column=0,columnspan=3,
            ipadx=10,ipady=10,
            padx=10,pady=10)

    def __loadFiles(self)->None:
        """
            Load purchase order documents from default/selected directory
        """
        self.__poFilesList = glob.glob(self.__srcDir + "/*.pdf") + glob.glob(self.__srcDir + "/*.txt")
        if len(self.__poFilesList)==0:
            self.__log(f"No pdf files found from directory {self.__srcDir}.")
        else:
            self.__log(f"{len(self.__poFilesList)} pdf files loaded from directory {self.__srcDir}.")

    def __selectDirectory(self)->None:
        """
            Select a source directory path
        """
        self.__srcDir = filedialog.askdirectory()
        self.__loadFiles()

    def __getPurchaseOrderFileType(self,poDocFilepath:str)->list:
        """
            Returns the buyer type of the purchase order file
        """
        poDoc = po_base.PO_BASE(poDocFilepath)
        if "PRIMARK" in poDoc.getPage(1).upper():
            return ["TYPE-1",None]
        elif "KMART" in poDoc.getPage(1).upper():
            return ["TYPE-2",None]
        elif "JUST GROUP" in poDoc.getPage(1).upper():
            return ["TYPE-4",None]
        elif "SALLING GROUP" in poDoc.getPage(1).upper():
            return ["TYPE-5",None]
        elif "TOKMANNI OY" in poDoc.getPage(1).upper():
            return ["TYPE-6",None]
        elif "PO BOX 85 WESTGATE NSW 2048" in poDoc.getPage(1).upper(): # Best & Less
            return ["TYPE-7",None]
        elif "RENFOLD LTD" in poDoc.getPage(1).upper():
            return ["TYPE-8",None]
        elif poDoc.getPage(1).upper()=="": 
            poDocContent = text_extract.extract(poDocFilepath,1,poDoc.numPages(),2.2)
            if "WAREHOUSE" in poDocContent[0].upper():
                return ["TYPE-3",poDocContent]

    def __extractData(self,poDocFilepath:str)->tuple:
        """
            Returns the extracted data
        """
        [poFileType,poDocContent] = self.__getPurchaseOrderFileType(poDocFilepath)
        match poFileType:
            case "TYPE-1":
                poDoc = po_type_1.PO_TYPE_1(poDocFilepath)
            case "TYPE-2":
                poDoc = po_type_2.PO_TYPE_2(poDocFilepath)
            case "TYPE-3":
                poDoc = po_type_3.PO_TYPE_3(poDocFilepath,poDocContent,2.2)
            case "TYPE-4":
                poDoc = po_type_4.PO_TYPE_4(poDocFilepath)
            case "TYPE-5":
                poDoc = po_type_5.PO_TYPE_5(poDocFilepath)
            case "TYPE-6":
                poDoc = po_type_6.PO_TYPE_6(poDocFilepath)
            case "TYPE-7":
                poDoc = po_type_7.PO_TYPE_7(poDocFilepath)
            case "TYPE-8":
                poDoc = po_type_8.PO_TYPE_8(poDocFilepath)                   
        (poDetails) = poDoc.output()
        return (poDetails, poFileType)

    def __startExtraction(self)->None:
        """
            Start the data extraction
        """
        self.__buttonSelect.configure(state='disabled')
        self.__buttonExtract.configure(state='disabled')

        if len(self.__poFilesList)>0:
            if not os.path.exists(self.__srcDir + "/completed"):
                os.mkdir(self.__srcDir + "/completed")
                self.__log("./completed dir created.")

            self.__log("Data extraction started.")
            for poDoc in self.__poFilesList:
                self.__log(f"[{os.path.basename(poDoc)}] Extraction started.")
                (poDetails, poFileType) = self.__extractData(poDoc)
                self.__log(f"[{os.path.basename(poDoc)}] Extraction completed.")
                self.__log(f"[{os.path.basename(poDoc)}] Writing data to excel.")
                self.__writeData(poDetails, poFileType)
                self.__log(f"[{os.path.basename(poDoc)}] Writing data to excel completed.")
                shutil.copy(poDoc,f"{self.__srcDir}/completed/")
                self.__log(f"[{os.path.basename(poDoc)}] moved to ./completed.")
                os.remove(poDoc)
            self.__log("Data extraction finished.")
        else:
            self.__log("No pdf files to extract.")
        self.__buttonSelect.configure(state='normal')
        self.__buttonExtract.configure(state='normal')

    def __writeData(self,poDetails:list,poFileType:str)->None:
        """
            Write extracted data to the standard excel format
        """
        try:
            buyer = poDetails[0]["buyer"].replace("THE ","").split(" ")[0]
        except IndexError:
            buyer = "_"

        excelFile = f'{self.__srcDir}/PO_{buyer}.xlsx'
        try:
            workbook = load_workbook(excelFile)
        except FileNotFoundError:
            shutil.copy("./po_std_format.xlsx",excelFile)
            workbook = load_workbook(excelFile)

        worksheet = workbook['1']
        maxRow = worksheet.max_row+1
        for poDetail in poDetails:
            purchaseOrders = poDetail['purchase_orders']
            for destNumber in purchaseOrders.keys():
                packsData = purchaseOrders[destNumber]['packs_data']
                for packData in packsData:
                    worksheet[f'D{maxRow}'].value = poDetail['company']
                    worksheet[f'E{maxRow}'].value = poDetail['consignee']
                    worksheet[f'F{maxRow}'].value = poDetail['buyer']
                    worksheet[f'I{maxRow}'].value = poDetail['season_year']
                    worksheet[f'J{maxRow}'].value = poDetail['season']
                    worksheet[f'K{maxRow}'].value = poDetail['style']
                    worksheet[f'L{maxRow}'].value = poDetail['style_desc']
                    worksheet[f'M{maxRow}'].value = poDetail['gmt_item']
                    worksheet[f'N{maxRow}'].value = poDetail['uom']
                    worksheet[f'P{maxRow}'].value = poDetail['total_qty']
                    worksheet[f'Q{maxRow}'].value = poDetail['currency']
                    worksheet[f'R{maxRow}'].value = poDetail['factory']
                    worksheet[f'S{maxRow}'].value = purchaseOrders[destNumber]['ship_date']
                    worksheet[f'W{maxRow}'].value = poDetail['fabric']
                    worksheet[f'Z{maxRow}'].value = purchaseOrders[destNumber]['size_range']
                    worksheet[f'AA{maxRow}'].value = purchaseOrders[destNumber]['dest_num']
                    if poFileType=="TYPE-1":
                        worksheet[f'AA{maxRow}'].value = purchaseOrders[destNumber]['dest_num']
                    else:
                        worksheet[f'AA{maxRow}'].value = poDetail['po_num']
                    worksheet[f'AB{maxRow}'].value = poDetail['po_date']
                    worksheet[f'AB{maxRow}'].value = poDetail['shipment_mode']
                    worksheet[f'AE{maxRow}'].value = purchaseOrders[destNumber]['dest']

                    worksheet[f'AF{maxRow}'].value = packData['pack_colour']
                    worksheet[f'AG{maxRow}'].value = packData['pack_sizes']
                    worksheet[f'AI{maxRow}'].value = packData['n_units']
                    worksheet[f'AJ{maxRow}'].value = packData['supplier_cost']
                    if poFileType!="TYPE-1" and poFileType!="TYPE-8":
                        worksheet[f'AK{maxRow}'].value = packData['supplier_cost']                   
                    maxRow +=1
        workbook.save(excelFile)
        workbook.close()

    def __saveLog(self)->None:
        os.makedirs(f"{self.__srcDir}/log",exist_ok=True)
        log_filename = f'{self.__srcDir}\log\po-extractor log {datetime.datetime.now().strftime("%d-%m-%Y %H-%M-%S")} .txt'
        with open(log_filename,'w') as log_file:
            log_file.write(self.__textFieldLog.get("1.0",END))
        self.__log(f'Log file saved to {os.getcwd()}')

    def __log(self, message:str)->None:
        """
            Create the log message with event details
        """
        self.__textFieldLog.configure(state='normal')
        self.__textFieldLog.insert(END,f'{datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | {message}\n')
        self.__textFieldLog.see(END) #auto scroll text field to the end
        self.__textFieldLog.configure(state='disabled')

    def run(self)->None:
        """
            Run the application
        """
        self.__log('PO Extractor started.')
        self.__loadFiles()
        self.__root.mainloop()

if __name__ == "__main__":
    app = PO_Extractor()
    app.run()